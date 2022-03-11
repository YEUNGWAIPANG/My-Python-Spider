import os
import re
import sys
import time
import ctypes
import requests
import openpyxl
import xlsxwriter
import pandas as pd
from lxml import etree
from base64 import b64encode
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QThread,pyqtSignal
from PyQt5.QtWidgets import QMainWindow,QApplication,QMessageBox

#获取课表类
class EduLessonTableSpider:
    def __init__(self,Account,Password) -> None:
        self.account = Account
        self.password = Password
        if not os.path.exists("./课表"):
            os.makedirs("./课表")
        self.session = requests.Session()
        # 学校教务系统登录时候要带Cookie,对首页发送请求时候获取到的Cookie提取出JSESSIONID并与browserID拼接成的才是真正有效的cookie，可以用来登录，也可以用来进行后面的操作。
        self.cookies = re.findall(r'(JSESSIONID=.*?) for',str(self.session.get('https://jwxt.gdupt.edu.cn/').cookies))[0] + '; browserID=3969211827'
        self.UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36 Edg/92.0.902.67'
        self.Host = 'jwxt.gdupt.edu.cn'
        self.termcode = self.__GetTermCode()
        self.week = None

    # 登录教务系统
    def Login(self) -> bool:
        headers = {
            'Host': self.Host,
            'Referer': 'https://jwxt.gdupt.edu.cn/',
            'User-Agent': self.UA
        }
        Logindata = {
            'Cookie':self.cookies,
            'account': self.account,
            # 学校教务系统发送用户信息时候密码会经base64加密再发送
            'pwd': str(b64encode(self.password.encode("utf-8"))).replace("b'","").replace("'",""),
            'verifycode':'' 
        }
        r = self.session.post('https://jwxt.gdupt.edu.cn/login!doLogin.action',headers=headers,data=Logindata)
        r = eval(r.text)
        if(r["msg"] == "/login!welcome.action"):
            result = True
        else:
            result = False
        return result

    # 获取teamcode供查询课表用
    def __GetTermCode(self) -> str:
        month = datetime.now().month
        if month in [8,9,10,11,12,1,2]:
            termcode = str(datetime.now().year) + "01"
        else:
            termcode = str(datetime.now().year - 1) + "02"
        return termcode

    # 从校历页面推测出当前是第几周
    def GetWeek(self) -> int:
        headers = {
            'Cookie':self.cookies,
            'Host': self.Host,
            'Referer': 'https://jwxt.gdupt.edu.cn/',
            'User-Agent': self.UA
        }
        params = {
            "xnxqdm":self.termcode,
            # 时间戳小数点前9位加小数点后3位
            "_":re.findall(r'(.*?)\.',str(time.time()))[0] + re.findall(r'.*?\.(\d{3})',str(time.time()))[0]
        }
        try:
            Week = int(etree.HTML(self.session.get("https://jwxt.gdupt.edu.cn/xlxx!getXlxx.action",headers=headers,params=params).text).xpath('//tr/td[text()="' + str(datetime.now().year) + "-" + str(datetime.now().month).zfill(2) +'"]/following-sibling::td[text()="' + str(datetime.now().day) +'"]/preceding-sibling::td[not(@*)]/text()')[0])
        except:
            print("本周未在本学期校历中.")
            Week = 0
        return Week

    # 检查当前学期的课程是否存在
    def CheckLesson(self) -> bool:
        headers = {
            'Host': self.Host,
            'Referer': 'https://jwxt.gdupt.edu.cn/xsgrkbcx!getXsgrbkList.action',
            'User-Agent': self.UA,
        }
        r = self.session.get('https://jwxt.gdupt.edu.cn/xsgrkbcx!xsAllKbList.action',headers=headers,params={'xnxqdm': self.termcode})
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        r = etree.HTML(r.text)
        if(len(r.xpath('//body/table')) != 0):
            result = True
        else:
            result = False
        return result

    # 获取课表
    def Getlesson(self) -> tuple:
        Lessons = []
        headers = {
            'Cookie':self.cookies,
            'Host': self.Host,
            'Referer': 'https://jwxt.gdupt.edu.cn/xsgrkbcx!xskbList.action?xnxqdm='+ str(self.termcode) + '&zc=' + str(self.week),
            'User-Agent': self.UA,
        }
        r = requests.get('https://jwxt.gdupt.edu.cn/xsgrkbcx!getKbRq.action',headers=headers,params={'xnxqdm': self.termcode,'zc': self.week})
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        Cdict = eval(r.text)
        LessonList = Cdict[0] #课程信息列表
        TimeList = Cdict[1] # 这一周的所有天以及对应的日期
        if(len(LessonList) == 0):
            print("\n本星期暂无课程!\n")
        else:  
            for Lesson in LessonList:
                LessonName = Lesson['kcmc']#课程名称
                LessonTeacher = Lesson['teaxms']#老师
                LessonDate = Lesson['xq']#星期几
                LessonPlace = Lesson['jxcdmc']#地点
                LessonJC = Lesson['jcdm2']#第几节
                Lessons.append((LessonName,LessonTeacher,LessonDate,LessonPlace,LessonJC))
        return Lessons,TimeList

# 开QThread执行检查教务系统是否开放的任务以免ui卡死
class CheckOpen(QThread):
    result = pyqtSignal(bool)

    def __init__(self) -> None:
        super(CheckOpen,self).__init__()
    
    def run(self):
        try:
            requests.get("https://jwxt.gdupt.edu.cn/")
            self.result.emit(True)
        except:
            self.result.emit(False)

# 开QThread执行获取课表的任务以免ui卡死
class GetLessonThread(QThread):
    info = pyqtSignal(str)

    def __init__(self,account,password):
        super(GetLessonThread,self).__init__()
        self.spider = EduLessonTableSpider(account,password)

    def save(self,TimeList,Lessons,TermCode,Week):
        TermCode = str(TermCode)
        Name = TermCode[0:4] + "学年第" + str(int(TermCode[4:6])) + "学期 第"+ str(Week) + "周课表.xlsx"
        workbook = xlsxwriter.Workbook("./课表/" + Name)
        worksheet = workbook.add_worksheet()
        rowhead = []
        Weekdays = ["星期日","星期一","星期二","星期三","星期四","星期五","星期六"]
        num = 7
        for day in Weekdays:
            for xq in TimeList:
                if int(xq["xqmc"]) == num:
                    str1 = day + " " + xq["rq"]
                    rowhead.append(str1)
                    if num == 7:
                        num = 1
                    else:
                        num += 1
                    break
        cowhead = ["第1节 8:00-8:45","第2节 8:55-9:40","第3节 10:00-10:45","第4节 10:55-11:40","第5节 14:30-15:15","第6节 15:25-16:10","第7节 16:20-16:55","第8节 17:05-17:50","第9节 19:40-20:25","第10节 20:35-21:20","第11节 21:30-22:15","第12节 22:25"]
        for i in rowhead:
            worksheet.write(0,rowhead.index(i)+1,i)
        for j in cowhead:
            if cowhead.index(j) > 3:
                if cowhead.index(j) > 7:
                    worksheet.write(cowhead.index(j)+3,0,j)
                else:
                    worksheet.write(cowhead.index(j)+2,0,j)
            else:
                worksheet.write(cowhead.index(j)+1,0,j)
        for lesson in Lessons:
            lessoninfo = lesson[0] + " " + lesson[1] + " " + lesson[3]
            lessontime = lesson[4].split(",")
            if int(lesson[2]) == 7:
                if int(lessontime[0]) > 4:
                    if int(lessontime[0]) > 8:
                        worksheet.write(int(lessontime[0])+2,1,lessoninfo)
                        worksheet.write(int(lessontime[1])+2,1,lessoninfo)
                    else:
                        worksheet.write(int(lessontime[0])+1,1,lessoninfo)
                        worksheet.write(int(lessontime[1])+1,1,lessoninfo)
                else:
                   worksheet.write(int(lessontime[0]),1,lessoninfo)
                   worksheet.write(int(lessontime[1]),1,lessoninfo)                         
            else:
                if int(lessontime[0]) > 4:
                    if int(lessontime[0]) > 8:
                        worksheet.write(int(lessontime[0])+2,int(lesson[2])+1,lessoninfo)
                        worksheet.write(int(lessontime[1])+2,int(lesson[2])+1,lessoninfo)
                    else:
                        worksheet.write(int(lessontime[0])+1,int(lesson[2])+1,lessoninfo)
                        worksheet.write(int(lessontime[1])+1,int(lesson[2])+1,lessoninfo)
                else:
                    worksheet.write(int(lessontime[0]),int(lesson[2])+1,lessoninfo)
                    worksheet.write(int(lessontime[1]),int(lesson[2])+1,lessoninfo)
        workbook.close()

    # 优化Excel文件格式，使单元格列宽适应
    def Pretty_excel(self,filename,sheetname) -> None:
        wb = openpyxl.load_workbook(filename)
        ws = wb["Sheet1"]

        df=pd.read_excel(filename,sheetname).fillna('-')
        df.loc[len(df)]=list(df.columns)						#把标题行附加到最后一行
        for col in df.columns:				
            index=list(df.columns).index(col)					#列序号
            letter=get_column_letter(index+1)					#列字母
            collen=df[col].apply(lambda x:len(str(x).encode())).max()	#获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width=collen*1.0+2	#也就是列宽为最大长度*1.2 可以自己调整
        
        #行和列表头居中
        for c in range(1,ws.max_column + 1):
            for r in range(1,ws.max_row + 1):
                ws.cell(row=r,column=c).alignment = Alignment(horizontal='center', vertical='center') 
        wb.save(filename)

    def run(self):
        if self.spider.Login():# 登陆成功
            self.info.emit("\n教务系统登陆成功！\n") 
            if self.spider.CheckLesson(): # 检查本学期的课程是否已经上线教务系统
                self.spider.week = self.spider.GetWeek()
                Name = self.spider.termcode[0:4] + "学年第" + str(int(self.spider.termcode[4:6])) + "学期 第"+ str(self.spider.week) + "周课表.xlsx"
                if not os.path.exists("./课表/" + Name):
                    info = self.spider.Getlesson()
                    if len(info) != 0:
                        lessons = info[0]
                        timelist = info[1]
                        self.info.emit("\n正在导出课表.\n")
                        self.save(timelist,lessons,self.spider.termcode,self.spider.week)
                        self.info.emit("\n课表导出成功.\n")
                        self.info.emit("\n正在优化课表格式。\n")
                        self.Pretty_excel("./课表/"+Name,"Sheet1")
                        self.info.emit("\n课表格式优化完成。\n")
                        
                    else:
                        self.info.emit("\n本周无课程。\n")
                else:
                    self.info.emit("\n本地存在该周课表.\n")
            else:
                self.info.emit("\n教务系统暂无本学期课程。\n")
        else:
            self.info.emit("\n用户名或密码错误,登陆失败。\n")
 
# 窗口主类
class Ui_MainWindow(object):

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(660, 600)
        MainWindow.setFixedSize(660, 600)
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("myappid")
        MainWindow.setWindowIcon(QtGui.QIcon("computer.ico"))
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.textEdit = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit.setGeometry(QtCore.QRect(20, 190, 621, 351))
        self.textEdit.setReadOnly(True)
        self.textEdit.setObjectName("textEdit")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(360, 120, 261, 34))
        self.pushButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.pushButton.setObjectName("pushButton")
        self.layoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.layoutWidget.setGeometry(QtCore.QRect(50, 50, 251, 111))
        self.layoutWidget.setObjectName("layoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.layoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label = QtWidgets.QLabel(self.layoutWidget)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)
        self.StuNumEntry = QtWidgets.QLineEdit(self.layoutWidget)
        self.StuNumEntry.setObjectName("StuNumEntry")
        self.horizontalLayout.addWidget(self.StuNumEntry)
        self.verticalLayout.addLayout(self.horizontalLayout)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.label_2 = QtWidgets.QLabel(self.layoutWidget)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_2.addWidget(self.label_2)
        self.PasswordEntry = QtWidgets.QLineEdit(self.layoutWidget)
        self.PasswordEntry.setObjectName("PasswordEntry")
        self.PasswordEntry.setEchoMode(QtWidgets.QLineEdit.Password)
        self.horizontalLayout_2.addWidget(self.PasswordEntry)
        self.verticalLayout.addLayout(self.horizontalLayout_2)
        self.RestartButton = QtWidgets.QPushButton(self.centralwidget)
        self.RestartButton.setGeometry(QtCore.QRect(360, 60, 261, 34))
        self.RestartButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.RestartButton.setObjectName("RestartButton")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 660, 30))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        MainWindow.setTabOrder(self.StuNumEntry, self.PasswordEntry)
        MainWindow.setTabOrder(self.PasswordEntry, self.pushButton)
        MainWindow.setTabOrder(self.pushButton, self.textEdit)

        self.StuNumEntry.setEnabled(False)
        self.PasswordEntry.setEnabled(False)
        self.pushButton.setEnabled(False)
        self.RestartButton.setEnabled(False)
        self.RestartButton.clicked.connect(self.check_system_open)
        self.pushButton.clicked.connect(self.getlesson_main)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "教务系统当周课表导出"))
        self.pushButton.setText(_translate("MainWindow", "登 录 并 导 出 当 周 课 表"))
        self.label.setText(_translate("MainWindow", "学 号"))
        self.label_2.setText(_translate("MainWindow", "密 码"))
        self.RestartButton.setText(_translate("MainWindow", "重 新 启 动"))

    def check_system_open(self):
        def getresult(result:bool):
            if result == True:
                self.textEdit.append("教务系统开放,请登录。\n")
                self.StuNumEntry.setEnabled(True)
                self.PasswordEntry.setEnabled(True)
                self.pushButton.setEnabled(True)
            else:
                self.textEdit.append("教务系统不在开放时段内或设备未联网,请于开放时段使用或检查网络状况后点击重新启动。\n")
                self.RestartButton.setEnabled(True)
        self.StuNumEntry.setEnabled(False)
        self.PasswordEntry.setEnabled(False)
        self.pushButton.setEnabled(False)
        self.RestartButton.setEnabled(False)
        self.checkthread = CheckOpen()
        self.checkthread.result.connect(getresult)
        self.checkthread.start()

    def getlesson_main(self):
        def PrintInfo(text:str):
            self.textEdit.append(text)
        def start_get():
            self.pushButton.setEnabled(False)
        def end_get():
            self.pushButton.setEnabled(True)
            self.textEdit.append("\n导出任务完成。\n")
        usernum = self.StuNumEntry.text()
        password = self.PasswordEntry.text()
        if usernum == "":
            QMessageBox.about(None,"错误","请输入学号")
        else:
            if password == "":
                QMessageBox.about(None,"错误","请输入密码")
            else:
                self.getlesson_thread = GetLessonThread(usernum,password)
                self.getlesson_thread.info.connect(PrintInfo)
                self.getlesson_thread.started.connect(start_get)
                self.getlesson_thread.finished.connect(end_get)
                self.getlesson_thread.start()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    MainWin = QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWin)
    MainWin.show()
    ui.check_system_open()
    sys.exit(app.exec_())
